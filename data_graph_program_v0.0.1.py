from tkinter import filedialog as F
from tkinter import messagebox as M
from tkinter import ttk
from tkinter import *
import time

import os

try:
	from matplotlib import pyplot as graph
except:
	os.system('python3 pip -m install matplotlib')
	from matplotlib import pyplot as graph

try:
	import numpy as ny
except:
	os.system('python3 pip -m install numpy')
	import numpy as ny
	
try:
	from openpyxl.utils import column_index_from_string
except:
	os.system('python3 pip -m install openpyxl')
	from openpyxl.utils import column_index_from_string

#from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter

#import numpy
###	a = numpy.asarray([ [1,2,3], [4,5,6], [7,8,9] ])
###	ny.savetxt("foo.csv", a, delimiter=",")

class App(Frame):
	
	def __init__(self, master):
	
		Frame.__init__(self, master, width = 1000, height = 1000)
		self.parent = master
		self.grid()
		self.parent.title("Excel Scrapper")

		self.GUI()
		
	def GUI(self):
	
		bl0 = ttk.Label(self)
		bl0.grid(row = 0, column = 0)
		
		bl1 = ttk.Label(self)
		bl1.grid(row = 4, column = 8)
		
		l0 = ttk.Label(self, text = 'Column Select:')
		l0.grid(row = 2, column = 1)
		
		self.e0 = ttk.Entry(self, width = 80)
		self.e0.grid(row = 1, column = 2, columnspan = 5, sticky = EW)
		
		self.e1 = ttk.Entry(self, width = 80)
		self.e1.grid(row = 2, column = 2, columnspan = 5, sticky = EW)
		
		self.b0 = ttk.Button(self, text = 'File:', command = self.filedial)
		self.b0.grid(row = 1, column = 1, sticky = EW)
		
		self.b1 = ttk.Button(self, text = 'Enter', command = self.process)
		self.b1.grid(row = 3, column = 1, columnspan = 6, sticky = EW)
		
	def filedial(self):
	
		self.filepath = F.askopenfilename(initialdir = 'desktop')
		self.e0.delete(0, 'end')
		self.e0.insert(0, self.filepath)
	
	def process(self):
		__cols_collected_flag__ = 0
		__file_opened_flag__ = 0
		__file_read_flag__ = 0
		try:		
			if self.filename:
				pass
		except:
			self.filename = self.e0.get()
		
		self.filesavedirc = os.path.dirname(self.filename)
		
		self.filesavedirc = self.filesavedirc+'/Graph Repository'
			
		try:
			os.makedirs(self.filesavedirc)
		except FileExistsError:
			pass
		
		#wb = openpyxl.load_workbook('origin.xlsx')
		#first_sheet = wb.get_sheet_names()[0]
		#worksheet = wb.get_sheet_by_name(first_sheet)
		
		'''workbook = load_workbook(self.filename)
		sheet = workbook.active
				
		cols = [int(i) for i in self.e1.get().split(',')]
		
		data = [ny.array([item.value for item in sheet[get_column_letter(i)]]) for i in cols]
		
		workbook.close()'''
		__key_words__ = ['time', 'TIME', 'Time']
		try:
			file = open(self.filename, 'r')
			__file_opened_flag__ = 1
		except FileNotFoundError:
			M.showerror(message = 'File Not Found Error:\n"'+self.filename+'"\nNot found!', title = 'Error!')
				    
		if __file_opened_flag__:
			try:
				raw_data = ny.array([line.split(',') for line in file])
				__file_read_flag__ = 1
			except:
				M.showerror(message = 'File Read Error:\n"'+self.filename+'"\nCould not opened!\n\nDid you use the wrong type of file?', title = 'Error!')
			if __file_read_flag__:
				file.close()
				ind = [[], []]
				[[[ind[0].append(item0) for item0 in line[0]], [ind[1].append(item1) for item1 in line[1]]] for line in [ny.where(ny.core.defchararray.find(raw_data, str)==0) for str in __key_words__]]
				__srow__ = ind[0][0]
				cols_data = self.e1.get().split(',')
				try:
					try:	
						__cols__ = [int(i) for i in cols_data]
					except ValueError:
						__cols__ = []
						for item in cols_data:
							try:
								__cols__.append(int(item))
							except ValueError:
								__cols__.append(column_index_from_string(item))
					__cols_collected_flag__ = 1
				except:
					M.showerror(message = 'Column Information Error:\nCould not understand column selection information!', title = 'Error!')

				if __cols_collected_flag__:
					labels = [raw_data[__srow__][i] for i in __cols__]
					__cols__.insert(0, ind[1][0])
					proc_data = ny.array([[float(item) for item in line] for line in raw_data[__srow__+1:, __cols__]])
					proc_data = proc_data.T

					self.fig_group = []
					for i in range(len(__cols__)-1):
						self.fig_group.append(graph.figure(i))
						current_figure = self.fig_group[-1].add_subplot(111)
						current_figure.plot(proc_data[0], proc_data[i])
						graph.grid()

						current_figure.set_title("Thermocouple Data "+labels[i])
						current_figure.set_xlabel("Time [min]")
						current_figure.set_ylabel("Temp. ["+u'\N{DEGREE SIGN}'+"C]")

						time_str = time.strftime('%Y_%m_%d-%H_%M_%S')

						self.fig_group[-1].savefig(self.filesavedirc+'/'+labels[i]+'_'+time_str+'.png', format = 'png')

					M.showinfo(message = 'Graphs Generated Successfully!\nFind images stored in:\n'+self.filesavedirc, title = 'Process Complete!')
		
root = Tk()
root.resizable(width = False, height = False)
app = App(root)
app.mainloop()
